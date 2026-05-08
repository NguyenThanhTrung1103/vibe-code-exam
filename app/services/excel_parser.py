"""Phase 05 — read-only streaming Excel parser.

Uses openpyxl `read_only=True, data_only=True` so cells are values not
formulas, and rows are streamed (no full-workbook in RAM).

Caller passes a workbook path + the column mapping
(`header_label -> canonical_field`). We yield one dict per data row.
"""

from __future__ import annotations

import unicodedata
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Canonical fields the rest of the pipeline understands.
# Phase 13 (CDEA Sprint-1) added: discussion_url, external_question_id,
# discussion_count, vote_a..vote_f. All optional — existing imports without
# any community columns continue to work unchanged.
CANONICAL_FIELDS = (
    "question_text",
    "question_type",
    "difficulty",
    "topic",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "option_e",
    "option_f",
    "option_g",
    "option_h",
    # `combined_options` — alternative to mapping option_a/b/.../h individually.
    # When mapped, the normalizer splits the cell value on `;` / `；` / newline
    # into option_a..option_h, so dump-style sheets with a single answer-list
    # column work end-to-end without admin pre-processing.
    "combined_options",
    "correct_answer",
    "explanation",
    "reference",
    "tags",
    # --- Phase 13 community-signal columns (all optional) ---
    "discussion_url",
    "external_question_id",
    "discussion_count",
    "vote_a",
    "vote_b",
    "vote_c",
    "vote_d",
    "vote_e",
    "vote_f",
    "vote_g",
    "vote_h",
)
REQUIRED_FIELDS = ("question_text", "option_a", "option_b", "correct_answer")
MAX_COLUMNS = 32  # safety cap — too-wide sheet → reject.


# Common header alias map for auto-mapping. Keys are the post-normalized
# (lowercase, accent-stripped, alphanumeric-only) form of the header label.
# Values are the canonical field names. Vietnamese aliases land here after
# accent stripping — see `_normalize_header()` below.
_ALIAS = {
    "question": "question_text",
    "questiontext": "question_text",
    "q": "question_text",
    "type": "question_type",
    "questiontype": "question_type",
    "level": "difficulty",
    "difficulty": "difficulty",
    "topic": "topic",
    "topics": "topic",
    "subject": "topic",
    "a": "option_a",
    "optiona": "option_a",
    "choicea": "option_a",
    "answera": "option_a",
    "b": "option_b",
    "optionb": "option_b",
    "choiceb": "option_b",
    "answerb": "option_b",
    "c": "option_c",
    "optionc": "option_c",
    "choicec": "option_c",
    "answerc": "option_c",
    "d": "option_d",
    "optiond": "option_d",
    "choiced": "option_d",
    "answerd": "option_d",
    "e": "option_e",
    "optione": "option_e",
    "choicee": "option_e",
    "answere": "option_e",
    "f": "option_f",
    "optionf": "option_f",
    "choicef": "option_f",
    "answerf": "option_f",
    "g": "option_g",
    "optiong": "option_g",
    "choiceg": "option_g",
    "answerg": "option_g",
    "h": "option_h",
    "optionh": "option_h",
    "choiceh": "option_h",
    "answerh": "option_h",
    "correct": "correct_answer",
    "correctanswer": "correct_answer",
    "answer": "correct_answer",
    "key": "correct_answer",
    "answerkey": "correct_answer",
    "explanation": "explanation",
    "rationale": "explanation",
    "reference": "reference",
    "url": "reference",
    "ref": "reference",
    "tags": "tags",
    "tag": "tags",
    # --- Phase 13 community-signal aliases ---
    "discussionurl": "discussion_url",
    "discussion": "discussion_url",
    "externalquestionid": "external_question_id",
    "extquestionid": "external_question_id",
    "extqid": "external_question_id",
    "discussioncount": "discussion_count",
    "comments": "discussion_count",
    "votea": "vote_a",
    "voteb": "vote_b",
    "votec": "vote_c",
    "voted": "vote_d",
    "votee": "vote_e",
    "votef": "vote_f",
    "voteg": "vote_g",
    "voteh": "vote_h",
    # --- Vietnamese / dump-style aliases (accent-stripped form) ---
    "cauhoi": "question_text",  # "Câu hỏi"
    "noidungcauhoi": "question_text",  # "Nội dung câu hỏi"
    # `Loại câu hỏi` literally means "question type". Without this entry the
    # `cauhoi` substring fallback would steal `question_text`.
    "loaicauhoi": "question_type",  # "Loại câu hỏi"
    "cautraloidung": "correct_answer",  # "Câu trả lời đúng"
    "dapandung": "correct_answer",  # "Đáp án đúng"
    "giaithichdapan": "explanation",  # "Giải thích đáp án" — canonical
    "giaithich": "explanation",  # "Giải thích"
    # "Mô tả thêm" used to alias to `explanation`, which collided with
    # "Giải thích đáp án" on Vietnamese dumps that carry both columns. Route
    # it to `reference` instead — its semantics ("additional notes") match
    # `reference` better and removes the duplicate-canonical mapping.
    "motathem": "reference",  # "Mô tả thêm"
    "chude": "topic",  # "Chủ đề"
    "linhvuc": "topic",  # "Lĩnh vực"
    "dokho": "difficulty",  # "Độ khó"
    "danhsachdapan": "combined_options",  # "Danh sách đáp án"
    "cacdapan": "combined_options",  # "Các đáp án"
    "dapan": "combined_options",  # "Đáp án" (alone — list)
    "thetag": "tags",  # "Thẻ tag"
}

# Sorted aliases (longest key first) so substring matching against long
# Vietnamese header labels prefers more specific matches.
_ALIAS_BY_LEN: list[tuple[str, str]] = sorted(_ALIAS.items(), key=lambda kv: -len(kv[0]))


@dataclass(slots=True)
class ParsedRow:
    sheet_name: str
    row_number: int  # 1-based, matches Excel row number
    raw: dict[str, Any]


def auto_map(headers: list[str]) -> dict[str, str | None]:
    """Best-effort header → canonical mapping. Unknown headers map to None.

    Lookup order per header:
      1. Exact match on the normalized header form.
      2. Substring match against any registered alias key (longest-first), so
         long descriptive labels like "Đáp án đúng (ví dụ trả lời 1...)" still
         resolve to `correct_answer` via the embedded "dapandung" substring.

    Duplicate-canonical de-duplication:
      If two headers both resolve to the same canonical field, keep the more
      specific match (exact alias wins over substring; longer alias key wins
      over shorter), and drop the rival back to None. Vietnamese sheets that
      carry both "Giải thích đáp án" and "Mô tả thêm" used to silently land
      on `explanation` twice; this guarantees a single owner per canonical.
    """
    out: dict[str, str | None] = {}
    # Specificity score per (header, canonical) so we can break ties later.
    # Higher score = more specific match.
    score: dict[str, int] = {}
    for h in headers:
        if not h:
            continue
        key = _normalize_header(h)
        if not key:
            out[h] = None
            continue
        match: str | None = _ALIAS.get(key)
        match_score = 0
        if match is not None:
            # Exact normalized match — strongest signal. Add the alias length
            # so e.g. "giaithichdapan" beats "giaithich" if both ever map to
            # the same canonical via different paths.
            match_score = 1000 + len(key)
        else:
            # Substring fallback only for aliases of length >= 4 — short aliases
            # ("a", "b", "q", "url"...) match too eagerly inside long words like
            # "random", "topic", "url" etc. Long Vietnamese alias keys
            # ("dapandung", "cauhoi", "danhsachdapan") still resolve.
            #
            # Two-pass: prefer aliases that appear at position 0 of the
            # header (the header *starts with* the alias) over aliases buried
            # later. Without this, `Tags (Các đáp án ...)` would lose `tags`
            # (pos 0, len 4) to `cacdapan` (pos 4, len 8) because the
            # length-desc loop saw `cacdapan` first.
            prefix_match: tuple[str, str] | None = None
            inner_match: tuple[str, str] | None = None
            for alias, candidate in _ALIAS_BY_LEN:
                if len(alias) < 4:
                    continue
                pos = key.find(alias)
                if pos < 0:
                    continue
                if pos == 0:
                    if prefix_match is None:
                        prefix_match = (alias, candidate)
                elif inner_match is None:
                    inner_match = (alias, candidate)
            if prefix_match is not None:
                alias, match = prefix_match
                match_score = 500 + len(alias)
            elif inner_match is not None:
                alias, match = inner_match
                match_score = len(alias)
        out[h] = match
        score[h] = match_score

    # Resolve duplicate canonicals by preferring the highest-scoring header
    # for each canonical; reset losers to None so the operator can re-pick.
    by_canonical: dict[str, str] = {}  # canonical → winning header
    for h, owner_canonical in list(out.items()):
        if not owner_canonical:
            continue
        winner = by_canonical.get(owner_canonical)
        if winner is None or score.get(h, 0) > score.get(winner, 0):
            if winner is not None:
                out[winner] = None  # demote previous winner
            by_canonical[owner_canonical] = h
        else:
            out[h] = None  # demote this header
    return out


def _normalize_header(label: str) -> str:
    """Lower-case + strip diacritics + keep alphanumerics only.

    Accent stripping lets Vietnamese accented forms ("Câu hỏi") and their
    ASCII transliterations ("Cau hoi") share a single alias entry.

    Vietnamese `đ` / `Đ` (U+0111 / U+0110) have no NFKD decomposition, so
    they are explicitly folded to ASCII `d` after the lower-case pass.
    """
    decomposed = unicodedata.normalize("NFKD", label.lower())
    no_accents = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
    no_accents = no_accents.replace("đ", "d")  # already lower-cased above
    return "".join(ch for ch in no_accents if ch.isalnum())


def read_headers(path: Path | str) -> tuple[str, list[str]]:
    """Open workbook, return `(sheet_name, header_row_values)`.

    Reads only the first sheet's first row.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        headers: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for cell in row:
                headers.append("" if cell is None else str(cell).strip())
            break
        if len(headers) > MAX_COLUMNS:
            raise ValueError(f"too many columns ({len(headers)} > {MAX_COLUMNS})")
        return ws.title, headers
    finally:
        wb.close()


def stream_rows(
    path: Path | str,
    *,
    column_mapping: dict[str, str | None],
    max_rows: int,
) -> Iterator[ParsedRow]:
    """Yield one `ParsedRow` per non-empty data row.

    `column_mapping` keys are the workbook's *header labels*; values are the
    canonical field names (or None to skip the column).

    Stops + raises `ValueError` if more than `max_rows` data rows are seen.
    Empty rows (all cells None) are skipped silently.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        sheet = ws.title
        header_row: list[str] = []
        seen_data = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                header_row = ["" if c is None else str(c).strip() for c in row]
                continue
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue
            seen_data += 1
            if seen_data > max_rows:
                raise ValueError(f"too many data rows (>{max_rows})")
            raw: dict[str, Any] = {}
            for header, value in zip(header_row, row, strict=False):
                canonical = column_mapping.get(header) if header else None
                if not canonical:
                    continue
                raw[canonical] = value
            if not raw:
                continue
            yield ParsedRow(sheet_name=sheet, row_number=row_idx, raw=raw)
    finally:
        wb.close()
